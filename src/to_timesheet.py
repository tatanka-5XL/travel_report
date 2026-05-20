#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Filename: to_timesheet.py
Description: Parses waypoints in travel report json input and fills in relevant timesheet

Author: Tatanka5XL
Created: 2026-01-29
Last Modified: 2026-02-03
Version: 0.4 - added R&D minutes and overall R&D percent calculations
License: Proprietary
"""

from __future__ import annotations

from datetime import datetime, timedelta
import os
import json
from openpyxl import load_workbook


# ---------------------------
# Date helpers (MMDD -> DD/MM)
# ---------------------------

def mmdd_to_date(year: str, mmdd: str) -> datetime:
    return datetime.strptime(f"{year}{mmdd}", "%Y%m%d")


def mmdd_to_ddmm(year: str, mmdd: str) -> str:
    return mmdd_to_date(year, mmdd).strftime("%d/%m")


def hhmm_to_hh_colon_mm(hhmm: str) -> str:
    hhmm = str(hhmm).zfill(4)
    return f"{hhmm[:2]}:{hhmm[2:]}"


# ---------------------------
# Build timeline segments (COLLAPSE border-cross travel points)
# ---------------------------

def build_segments(data: dict) -> list[dict]:
    year = str(data["year"])
    out: list[dict] = []

    # Flatten all waypoints into a single chronological list to handle midnight/multi-day logic
    all_wps = []
    for mmdd in sorted(data["waypoints"].keys(), key=int):
        for wp in data["waypoints"][mmdd]:
            # Attach date info to the waypoint object for easier math
            wp_copy = wp.copy()
            wp_copy["_dt"] = datetime.strptime(f"{year}{mmdd}{str(wp['time']).zfill(4)}", "%Y%m%d%H%M")
            wp_copy["_mmdd"] = mmdd
            all_wps.append(wp_copy)

    i = 0
    while i < len(all_wps) - 1:
        cur = all_wps[i]
        nxt = all_wps[i + 1]
        kind = (cur.get("next") or "").strip().lower()

        # 1. Handle HOTEL (New type)
        if kind == "hotel":
            # Just skip to next; doesn't add minutes/km to timesheet
            i += 1
            continue

        # 2. Only process TRAVEL or MEETING
        if kind not in ("travel", "meeting"):
            i += 1
            continue

        start_dt = cur["_dt"]
        end_dt = nxt["_dt"]

        # Ensure end is after start (handles basic cross-midnight logic)
        if end_dt < start_dt:
            end_dt = end_dt.replace(day=end_dt.day + 1)

        # Logic to split segments if they cross midnight
        temp_start = start_dt
        while temp_start < end_dt:
            # Determine the end of the current day
            day_end = temp_start.replace(hour=23, minute=59, second=59)
            # Use the actual arrival time or midnight, whichever comes first
            current_seg_end = min(end_dt, day_end)
            
            duration_mins = int(round((current_seg_end - temp_start).total_seconds() / 60.0))
            if duration_mins < 0: duration_mins = 0

            seg_mmdd = temp_start.strftime("%m%d")
            
            out.append({
                "mmdd": seg_mmdd,
                "date_out": temp_start.strftime("%d/%m"),
                "start_hhmm": temp_start.strftime("%H%M"),
                "end_hhmm": current_seg_end.strftime("%H%M"),
                "type": kind,
                "country": (cur.get("country") or "").strip().upper() if kind == "travel" else (cur.get("country") or "").strip().upper(),
                "place_from": cur.get("place") or "",
                "place_to": nxt.get("place") or "",
                "minutes": duration_mins,
                "km": int(nxt.get("km", 0) or 0) if temp_start == start_dt else 0, # Add KM only to the first part of split
                "rd": int(cur.get("r_d", 0) or 0) if kind == "meeting" else 0,
            })

            # Move start to the beginning of the next day
            if current_seg_end == day_end:
                temp_start = (day_end + timedelta(seconds=1)).replace(second=0)
            else:
                temp_start = end_dt

        i += 1

    return out


def find_first_meeting_start(segments: list[dict]) -> tuple[str, str] | None:
    for s in segments:
        if s["type"] == "meeting":
            return s["mmdd"], s["start_hhmm"]
    return None


def find_last_meeting_end(segments: list[dict]) -> tuple[str, str] | None:
    last = None
    for s in segments:
        if s["type"] == "meeting":
            last = (s["mmdd"], s["end_hhmm"])
    return last


# ---------------------------
# Fill the template
# ---------------------------

def fill_timesheet(template_path: str, data: dict, out_path: str) -> None:
    year = str(data["year"])
    segs = build_segments(data)
    if not segs:
        raise ValueError("No segments were built. Check your JSON waypoints / 'next' fields.")

    day_keys = sorted(data["waypoints"].keys(), key=int)
    trip_from = mmdd_to_ddmm(year, day_keys[0])
    trip_to = mmdd_to_ddmm(year, day_keys[-1])

    first_meet = find_first_meeting_start(segs)
    last_meet = find_last_meeting_end(segs)

    wb = load_workbook(template_path)
    ws = wb.active

    # Header
    ws["C7"].value = int(data["trip_info"]["trip_number"])
    ws["B5"].value = f"{trip_from} - {trip_to}"

    def key(mmdd: str, hhmm: str) -> int:
        return int(mmdd) * 10000 + int(hhmm)

    first_meet_key = key(*first_meet) if first_meet else None
    last_meet_key = key(*last_meet) if last_meet else None

    # Split travel there / home
    travel_there: list[dict] = []
    travel_home: list[dict] = []

    for s in segs:
        s_key_start = key(s["mmdd"], s["start_hhmm"])
        s_key_end = key(s["mmdd"], s["end_hhmm"])

        if s["type"] == "travel" and first_meet_key is not None and s_key_end <= first_meet_key:
            travel_there.append(s)

        if s["type"] == "travel" and last_meet_key is not None and s_key_start >= last_meet_key:
            travel_home.append(s)

    # Aggregate driving-only blocks -> one row per day
    def aggregate_daily(driving_segments: list[dict]) -> list[dict]:
        by_day: dict[str, list[dict]] = {}
        for s in driving_segments:
            by_day.setdefault(s["mmdd"], []).append(s)

        out: list[dict] = []
        for mmdd in sorted(by_day.keys(), key=int):
            items = sorted(by_day[mmdd], key=lambda x: int(x["start_hhmm"]))
            minutes = sum(int(x["minutes"]) for x in items)
            km = sum(int(x["km"]) for x in items)
            out.append({
                "mmdd": mmdd,
                "date_out": items[0]["date_out"],
                "start_hhmm": items[0]["start_hhmm"],
                "end_hhmm": items[-1]["end_hhmm"],
                "minutes": minutes,
                "km": km,
            })
        return out

    there_days = aggregate_daily(travel_there)
    home_days = aggregate_daily(travel_home)

    # first meeting place (for description)
    first_meeting_place = None
    for s in segs:
        if s["type"] == "meeting":
            first_meeting_place = s["place_from"]
            break

    # Remove travel_there/home segments from detailed
    used_keys = set((s["mmdd"], s["start_hhmm"], s["end_hhmm"], s["type"]) for s in (travel_there + travel_home))
    detailed = [
        s for s in segs
        if (s["mmdd"], s["start_hhmm"], s["end_hhmm"], s["type"]) not in used_keys
    ]

    # Sort detailed chronologically (important for "next meeting RD%" on drives)
    detailed.sort(key=lambda s: key(s["mmdd"], s["start_hhmm"]))

    # ---------------------------
    # IMPORTANT CHANGE #2:
    # For 2nd group TRAVEL segments, set rd = NEXT meeting rd%
    # ---------------------------
    next_meeting_rd = 0
    # Walk backwards so we always know "next meeting rd"
    for s in reversed(detailed):
        if s["type"] == "meeting":
            next_meeting_rd = int(s.get("rd", 0) or 0)
        elif s["type"] == "travel":
            s["rd"] = int(next_meeting_rd or 0)

    # ---------------------------
    # Row writer: sets E/F/G exactly as requested
    # ---------------------------
    def write_row(r: int, date_out: str, desc: str, start_hhmm: str, end_hhmm: str, rd_pct: float, minutes_total: int, km: int):
        minutes_total = int(minutes_total or 0)
        rd_pct = float(rd_pct or 0.0)
        rd_minutes = round(minutes_total * (rd_pct / 100.0), 2)

        ws.cell(row=r, column=1, value=date_out)                         # A
        ws.cell(row=r, column=2, value=desc)                             # B
        ws.cell(row=r, column=3, value=hhmm_to_hh_colon_mm(start_hhmm))   # C
        ws.cell(row=r, column=4, value=hhmm_to_hh_colon_mm(end_hhmm))     # D
        ws.cell(row=r, column=5, value=round(rd_pct, 2))                 # E = R&D %
        ws.cell(row=r, column=6, value=rd_minutes)                       # F = R&D minutes
        ws.cell(row=r, column=7, value=minutes_total)                    # G = duration minutes
        ws.cell(row=r, column=8, value=int(km or 0))                     # H = km

        return rd_minutes

    # ==========================================================
    # 1) SECOND GROUP FIRST: totals into C31, F31, G31
    # ==========================================================
    second_total_minutes = sum(int(s.get("minutes", 0) or 0) for s in detailed)
    second_total_rd_minutes = 0.0
    for s in detailed:
        m = int(s.get("minutes", 0) or 0)
        pct = float(s.get("rd", 0) or 0)
        second_total_rd_minutes += (m * pct / 100.0)

    second_total_rd_minutes = round(second_total_rd_minutes, 2)

    ws["G31"].value = int(second_total_minutes)
    ws["F31"].value = float(second_total_rd_minutes)

    if second_total_minutes > 0:
        avg_rd_pct = round((second_total_rd_minutes / second_total_minutes) * 100.0, 2)
    else:
        avg_rd_pct = 0.0

    ws["C31"].value = avg_rd_pct

    # ==========================================================
    # 2) WRITE FIRST GROUP: travel there + travel home
    #    Use avg_rd_pct in E, compute F from G
    # ==========================================================
    row = 10

    for d in there_days:
        write_row(
            r=row,
            date_out=d["date_out"],
            desc=f"Travel to {first_meeting_place or 'first meeting'}",
            start_hhmm=d["start_hhmm"],
            end_hhmm=d["end_hhmm"],
            rd_pct=avg_rd_pct,
            minutes_total=d["minutes"],
            km=d["km"],
        )
        row += 1

    for d in home_days:
        write_row(
            r=row,
            date_out=d["date_out"],
            desc="Travel home",
            start_hhmm=d["start_hhmm"],
            end_hhmm=d["end_hhmm"],
            rd_pct=avg_rd_pct,
            minutes_total=d["minutes"],
            km=d["km"],
        )
        row += 1

    # ==========================================================
    # TOTALS FOR BOTH GROUPS -> F33 (R&D minutes) and G33 (minutes)
    # Put this AFTER writing travel there/home rows and AFTER avg_rd_pct is known
    # ==========================================================

    # First group (travel there + travel home) totals
    first_total_minutes = sum(int(d.get("minutes", 0) or 0) for d in (there_days + home_days))
    first_total_rd_minutes = round(first_total_minutes * (avg_rd_pct / 100.0), 2)

    # Second group totals already computed above:
    # second_total_minutes
    # second_total_rd_minutes

    both_total_minutes = int(first_total_minutes) + int(second_total_minutes)
    both_total_rd_minutes = round(first_total_rd_minutes + float(second_total_rd_minutes), 2)

    ws["F33"].value = both_total_rd_minutes   # Total R&D minutes (both groups)
    ws["G33"].value = both_total_minutes      # Total minutes (both groups)    

    # Blank line
    row += 1

    # ==========================================================
    # 3) WRITE SECOND GROUP: meetings + travel between meetings
    #    E = s["rd"] (meeting’s own rd, travel uses next meeting rd),
    #    F computed, G minutes
    # ==========================================================
    for s in detailed:
        if s["type"] == "travel":
            desc = f"Travel to {s['place_to']} ({s['country']})"
        else:
            desc = f"Meeting at {s['place_from']} ({s['country']})"

        write_row(
            r=row,
            date_out=s["date_out"],
            desc=desc,
            start_hhmm=s["start_hhmm"],
            end_hhmm=s["end_hhmm"],
            rd_pct=float(s.get("rd", 0) or 0),
            minutes_total=int(s.get("minutes", 0) or 0),
            km=int(s.get("km", 0) or 0),
        )
        row += 1

    # Save
    out_path = os.path.expanduser(out_path)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


# ---------------------------
# CLI / usage
# ---------------------------
if __name__ == "__main__":
    default_json = "0101_itfr.json"
    json_filename = input(f"Input JSON [{default_json}]: ").strip() or default_json
    json_path = os.path.join("..", "input", json_filename)

    if not os.path.isfile(json_path):
        raise FileNotFoundError(f"Input JSON not found: {json_path}")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    template = os.path.join("..", "config", "timesheet_template.xlsx")
    if not os.path.isfile(template):
        raise FileNotFoundError(f"Template not found: {template}")

    out = os.path.expanduser(
        f"~/Documents/profi/expenses/timesheet_{data['report_id']}_{data['trip_info']['trip_number']}.xlsx"
    )

    fill_timesheet(template, data, out)
    print("Saved:", out)