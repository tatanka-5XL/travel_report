import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.pagebreak import Break


# ===========================
# Excel sheet generating part
# ===========================


wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)

# Print settings
sheet.print_area = "A1:F66"

# A4 + portrait (or switch to landscape if needed)
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
# try LANDSCAPE for wide tables
sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT

# Fit the print area to ONE page (this is the key)
sheet.page_setup.fitToWidth = 1
sheet.page_setup.fitToHeight = 1
sheet.sheet_properties.pageSetUpPr.fitToPage = True

# Reasonable margins (in inches!)
sheet.page_margins.left = 0.25
sheet.page_margins.right = 0.25
sheet.page_margins.top = 0.5
sheet.page_margins.bottom = 0.5

# Optional: center on page
sheet.page_setup.horizontalCentered = True


sheet.print_area = "A1:F85"
sheet.row_breaks.append(Break(id=66))
sheet.col_breaks.append(Break(id=6))

# Set default font for the workbook (matches Excel default)
default_font = Font(name="Helvetica", size=10)
for row in sheet.iter_rows(min_row=1, max_row=66, min_col=1, max_col=6):
    for cell in row:
        cell.font = default_font

for row_idx in range(1, 67):
    sheet.row_dimensions[row_idx].height = 12

sheet.column_dimensions["A"].width = 82 / 5.9
sheet.column_dimensions["B"].width = 266 / 5.9
sheet.column_dimensions["C"].width = 63 / 5.9
sheet.column_dimensions["D"].width = 63 / 5.9
sheet.column_dimensions["E"].width = 63 / 5.9
sheet.column_dimensions["F"].width = 63 / 5.9

# Text alignment in cells
alignment = Alignment(horizontal="left", vertical="center")

for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column
):
    for cell in row:
        cell.alignment = alignment

# Header
sheet['A1'] = 'Vyúčtování služební cesty'
sheet["A1"].font = Font(bold=True)

sheet['C1'] = 'Profisolv, s.r.o.'
sheet["C1"].font = Font(bold=True)

sheet['E1'] = 'Číslo:'
sheet["E1"].font = Font(bold=True)
sheet['E2'] = 'List:'
sheet["E2"].font = Font(bold=True)

sheet['A4'] = 'Pracovník:'
sheet["A4"].font = Font(bold=True)
sheet['A5'] = 'Ùčel cesty:'
sheet["A5"].font = Font(bold=True)
sheet['A6'] = 'Prostředek:'
sheet["A6"].font = Font(bold=True)
sheet['A7'] = 'Trasa:'
sheet["A7"].font = Font(bold=True)

# Route
sheet['B9'] = 'Popis trasy'
sheet["B9"].font = Font(bold=True)
sheet["B9"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

sheet['A10'] = 'Bod'
sheet["A10"].font = Font(bold=True)
sheet['B10'] = 'Místo'
sheet["B10"].font = Font(bold=True)
sheet['C10'] = 'Datum'
sheet["C10"].font = Font(bold=True)
sheet['D10'] = 'Čas'
sheet["D10"].font = Font(bold=True)
sheet['E10'] = 'Doba'
sheet["E10"].font = Font(bold=True)
sheet['F10'] = 'Jídla'
sheet["F10"].font = Font(bold=True)


# Costs
sheet['B32'] = 'Náklady'
sheet["B32"].font = Font(bold=True)
sheet["B32"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

sheet['A33'] = 'Stravné'
sheet["A33"].font = Font(bold=True)
sheet['A34'] = 'Datum'
sheet["A34"].font = Font(bold=True)
sheet['B34'] = 'Popis'
sheet["B34"].font = Font(bold=True)
sheet['E34'] = 'Plné'
sheet["E34"].font = Font(bold=True)
sheet['F34'] = 'Snížené'
sheet["F34"].font = Font(bold=True)
sheet['D42'] = 'Celkem:'
sheet['D43'] = 'Kapesné:'
sheet['F43'] = 'xxxxxxx'
sheet["F43"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

sheet['A45'] = 'Ubytování'
sheet["A45"].font = Font(bold=True)
sheet['A46'] = 'Datum'
sheet["A46"].font = Font(bold=True)
sheet['B46'] = 'Popis'
sheet["B46"].font = Font(bold=True)
sheet['E46'] = 'Doklad č.'
sheet["E46"].font = Font(bold=True)
sheet['F46'] = 'Částka'
sheet["F46"].font = Font(bold=True)
sheet['E52'] = 'Celkem:'

sheet['A54'] = 'Ostatní výdaje'
sheet["A54"].font = Font(bold=True)
sheet['A55'] = 'Datum'
sheet["A55"].font = Font(bold=True)
sheet['B55'] = 'Popis'
sheet["B55"].font = Font(bold=True)
sheet['E55'] = 'Doklad č.'
sheet["E55"].font = Font(bold=True)
sheet['F55'] = 'Částka'
sheet["F55"].font = Font(bold=True)
sheet['E62'] = 'Celkem:'

# Footer
sheet['A64'] = 'Zúčtováno dne:'
sheet['A65'] = 'Podpis'
sheet["E64"] = 'Záloha'
sheet['C65'] = 'Mezisoučet:'
sheet["C65"].font = Font(italic=True)
sheet["E65"] = 'Náklady'
sheet['E66'] = 'K vyplacení:'
sheet["E66"].font = Font(bold=True)

wb.save('../output/spam.xlsx')
