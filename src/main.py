#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Filename: main.py
Description: Parses segments in travel report json input and calculates times and perdiems.

Author: Tatanka5XL
Created: 2025-12-23
Last Modified: 2026-01-10
Version: 0.4
License: Proprietary
"""

from datetime import datetime
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.pagebreak import Break

# =========================
# Helpers
# =========================


def to_isotime(string_year, string_time):
    return datetime.strptime(f"{string_year}{string_time}", "%Y%m%d%H%M")


def to_stringday(isotime):
    return isotime.strftime("%d/%m")


def diff_in_hours(iso_start, iso_end):
    delta = iso_end - iso_start
    return delta.total_seconds() / 3600


def cz_band(hours: float):
    if hours < 5:
        return None
    if hours < 12:
        return "5_to_12"
    if hours < 18:
        return "12_to_18"
    return "over_18"


def foreign_band(hours: float):
    # entitlement bands for foreign: 1_to_12, 12_to_18, over_18
    if hours < 1:
        return None
    if hours < 12:
        return "1_to_12"
    if hours < 18:
        return "12_to_18"
    return "over_18"


def apply_meal_reduction(base_amount: float, pct_per_meal: float, meals: int) -> float:
    # reduce by pct_per_meal * meals, cap at 0
    factor = 1.0 - (meals * (pct_per_meal / 100.0))
    if factor < 0:
        factor = 0.0
    return base_amount * factor


# =========================
# Ask for input file
# =========================

default_file = "trep.json"
filename = input(
    f"Input JSON filename [{default_file}]: ").strip() or default_file

input_path = os.path.join("..", "input", filename)
if not os.path.isfile(input_path):
    raise FileNotFoundError(f"Input file not found: {input_path}")

with open(input_path, "r", encoding="utf-8") as f:
    json_file_data = json.load(f)

year = json_file_data["year"]


# =========================
# Build day overview (trip)
# =========================

trip = []

for segment in json_file_data["segments"]:
    start_time = to_isotime(year, segment["start_time"])
    end_time = to_isotime(year, segment["end_time"])
    day_str = to_stringday(start_time)
    hours = diff_in_hours(start_time, end_time)

    # create day if missing
    day = next((d for d in trip if d["date"] == day_str), None)
    if day is None:
        day = {"date": day_str, "segments": []}
        trip.append(day)

    # add or accumulate country segment
    seg = next((s for s in day["segments"]
               if s["country"] == segment["country"]), None)
    if seg is None:
        day["segments"].append({
            "country": segment["country"],
            "time_hours": hours,
            "meals": segment["meals"]
        })
    else:
        seg["time_hours"] += hours
        seg["meals"] += segment["meals"]


# =========================
# Load settings
# =========================

settings_path = os.path.join("..", "config", "settings.json")
if not os.path.isfile(settings_path):
    raise FileNotFoundError(f"Settings file not found: {settings_path}")

with open(settings_path, "r", encoding="utf-8") as f:
    settings = json.load(f)

cz_rates = settings["cz"]["per_diems_czk"]
cz_meal_reduce = settings["cz"]["lowering_percents_per_meal"]

foreign_rates_raw = settings["foreign"]["per_diems"]
foreign_percents = settings["foreign"]["per_diems_percents"]
foreign_meal_reduce = settings["foreign"]["lowering_percents_per_meal"]
pocket_percent = float(settings["foreign"].get("pocket_money_percent", 40))

# Build mapping: "PL" -> {"rate": 50, "currency": "EUR"} from keys like "PL_eur"
foreign_rates = {}
for k, v in foreign_rates_raw.items():
    country, cur = k.split("_", 1)
    foreign_rates[country] = {"rate": float(v), "currency": cur.upper()}

# Exchange rates STRICTLY from input JSON (CNB trip start day)
rates_czk = {}
for c in json_file_data["bank_rates"]["currencies"]:
    code = (c["code"] or "").upper().strip()
    rates_czk[code] = float(c["exchange_rate"])

if "CZK" not in rates_czk:
    raise KeyError(
        "bank_rates.currencies must include CZK with exchange_rate = 1")


# =========================
# Calculate per diems per day
# =========================

for day in trip:
    day["per_diem"] = []
    comments = []

    # ---------- CZ ----------
    cz_seg = next((s for s in day["segments"] if s["country"] == "CZ"), None)
    cz_hours = float(cz_seg.get("time_hours", 0) or 0) if cz_seg else 0.0
    cz_meals = int(cz_seg.get("meals", 0) or 0) if cz_seg else 0

    cz_key = cz_band(cz_hours)
    if cz_key is None:
        day["per_diem"].append({
            "country": "CZ",
            "currency": "CZK",
            "band": "under_5",
            "base": 0,
            "meals": cz_meals,
            "reduction_percent_per_meal": None,
            "amount": 0
        })
        if cz_seg:
            comments.append(f"CZ: {cz_hours:.2f}h <5h ⇒ 0 CZK.")
    else:
        cz_base = float(cz_rates[cz_key])
        cz_pct = float(cz_meal_reduce[cz_key])
        cz_final = apply_meal_reduction(cz_base, cz_pct, cz_meals)
        cz_lowered = cz_base - cz_final

        day["per_diem"].append({
            "country": "CZ",
            "currency": "CZK",
            "band": cz_key,
            "base": cz_base,
            "meals": cz_meals,
            "reduction_percent_per_meal": cz_pct,
            "amount": round(cz_final, 2)
        })
        comments.append(
            f"CZ: {cz_hours:.2f}h ⇒ base {cz_base:.2f} CZK ({cz_key}); "
            f"meals {cz_meals} lowered {cz_lowered:.2f} ⇒ paid {cz_final:.2f} CZK."
        )

    # ---------- FOREIGN ----------
    foreign_segs = [s for s in day["segments"] if s["country"] != "CZ"]
    if not foreign_segs:
        day["comment"] = " | ".join(comments)
        continue

    total_foreign_hours = sum(float(s.get("time_hours", 0) or 0)
                              for s in foreign_segs)
    total_foreign_meals = sum(int(s.get("meals", 0) or 0)
                              for s in foreign_segs)

    # Rule: if CZ >= 5h but foreign < 5h => no foreign per diem
    if cz_hours >= 5 and total_foreign_hours < 5:
        day["per_diem"].append({
            "country": "FOREIGN",
            "currency": None,
            "band": "blocked_cz>=5_foreign<5",
            "base_rate": None,
            "percent": 0,
            "base": 0,
            "foreign_hours_total": round(total_foreign_hours, 2),
            "meals": total_foreign_meals,
            "reduction_percent_per_meal": None,
            "amount": 0,
            "amount_czk": 0,
            "exchange_rate_to_czk": None
        })
        comments.append(
            f"FOREIGN: {total_foreign_hours:.2f}h, but CZ {cz_hours:.2f}h ≥5h and foreign <5h ⇒ 0."
        )
        day["comment"] = " | ".join(comments)
        continue

    # dominant foreign country by time (no FX logic for choosing)
    dominant_seg = max(foreign_segs, key=lambda s: float(
        s.get("time_hours", 0) or 0))
    country = dominant_seg["country"]

    if country not in foreign_rates:
        raise KeyError(
            f"No foreign per diem rate in settings.json for country: {country}")

    band_key = foreign_band(total_foreign_hours)
    cur = foreign_rates[country]["currency"]
    rate = foreign_rates[country]["rate"]

    if band_key is None:
        day["per_diem"].append({
            "country": country,
            "currency": cur,
            "band": "under_1",
            "base_rate": rate,
            "percent": 0,
            "base": 0,
            "foreign_hours_total": round(total_foreign_hours, 2),
            "meals": total_foreign_meals,
            "reduction_percent_per_meal": None,
            "amount": 0,
            "amount_czk": 0,
            "exchange_rate_to_czk": None
        })
        comments.append(
            f"FOREIGN ({country}): {total_foreign_hours:.2f}h <1h ⇒ 0.")
        day["comment"] = " | ".join(comments)
        continue

    percent = float(foreign_percents[band_key])
    foreign_base = rate * (percent / 100.0)

    red_pct = float(foreign_meal_reduce.get(band_key, 0))
    foreign_final = apply_meal_reduction(
        foreign_base, red_pct, total_foreign_meals)
    foreign_lowered = foreign_base - foreign_final

    foreign_amount = round(foreign_final, 2)

    if cur not in rates_czk:
        raise KeyError(
            f"Missing exchange rate for {cur} in input JSON bank_rates.currencies (CNB start-day)."
        )

    foreign_amount_czk = round(foreign_amount * rates_czk[cur], 2)

    day["per_diem"].append({
        "country": country,
        "currency": cur,
        "band": band_key,
        "base_rate": rate,
        "percent": percent,
        "base": round(foreign_base, 2),
        "foreign_hours_total": round(total_foreign_hours, 2),
        "meals": total_foreign_meals,
        "reduction_percent_per_meal": red_pct,
        "amount": foreign_amount,
        "exchange_rate_to_czk": rates_czk[cur],
        "amount_czk": foreign_amount_czk
    })

    comments.append(
        f"FOREIGN ({country}): total {total_foreign_hours:.2f}h ⇒ {band_key} ({percent:.1f}%) "
        f"base {foreign_base:.2f} {cur}; meals {total_foreign_meals} lowered {foreign_lowered:.2f} ⇒ "
        f"paid {foreign_final:.2f} {cur}. "
        f"Converted: {foreign_amount:.2f} {cur} × {rates_czk[cur]:.4f} = {foreign_amount_czk:.2f} CZK."
    )

    day["comment"] = " | ".join(comments)


# =========================
# Pocket money + totals (CZK)
#  - pocket money is % of per-diem BASES before meal reduction
#  - foreign bases converted to CZK using CNB start-day rates from input JSON
# =========================

per_diem_base_total_czk = 0.0
per_diem_paid_total_czk = 0.0

for day in trip:
    # Sum bases before reductions (convert foreign base to CZK)
    for p in day.get("per_diem", []):
        base = float(p.get("base", 0) or 0)
        cur = p.get("currency")

        # skip synthetic foreign placeholders with currency None
        if not cur:
            continue

        if cur == "CZK":
            per_diem_base_total_czk += base
        else:
            if cur not in rates_czk:
                raise KeyError(
                    f"Missing exchange rate for {cur} in input JSON bank_rates.currencies (CNB start-day)."
                )
            per_diem_base_total_czk += base * rates_czk[cur]

    # Sum actually paid per diems in CZK
    cz_paid = 0.0
    cz_entry = next((p for p in day.get("per_diem", [])
                    if p.get("country") == "CZ"), None)
    if cz_entry:
        cz_paid = float(cz_entry.get("amount", 0) or 0)

    foreign_paid_czk = 0.0
    foreign_entry = next((p for p in day.get("per_diem", []) if p.get(
        "currency") and p.get("currency") != "CZK"), None)
    if foreign_entry:
        foreign_paid_czk = float(foreign_entry.get("amount_czk", 0) or 0)

    per_diem_paid_total_czk += (cz_paid + foreign_paid_czk)

pocket_money_czk = round(per_diem_base_total_czk * (pocket_percent / 100.0), 2)

summary = {
    "total_per_diem_base_czk": round(per_diem_base_total_czk, 2),
    "total_per_diem_paid_czk": round(per_diem_paid_total_czk, 2),
    "pocket_money_percent": pocket_percent,
    "total_pocket_money_czk": pocket_money_czk,
    "total_money_czk": round(per_diem_paid_total_czk + pocket_money_czk, 2),
}


# =========================
# Output: days + summary
# =========================

base_name = os.path.splitext(filename)[0]
output_filename = f"{base_name}_seg.json"
output_path = os.path.join("..", "output", output_filename)

output = {
    "days": trip,
    "summary": summary
}

print(json.dumps(output, indent=2, ensure_ascii=False))

with open(output_path, "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"\nProcessed data saved to {output_path}")


# ===========================
# Excel sheet generating part
# ===========================


wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)

# Print settings
sheet.print_area = "A1:F66"

# A4 + portrait (or switch to landscape if needed)
sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
# try LANDSCAPE for wide tables
sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT

# Fit the print area to ONE page (this is the key)
sheet.page_setup.fitToWidth = 1
sheet.page_setup.fitToHeight = 1
sheet.sheet_properties.pageSetUpPr.fitToPage = True

# Reasonable margins (in inches!)
sheet.page_margins.left = 0.25
sheet.page_margins.right = 0.25
sheet.page_margins.top = 0.5
sheet.page_margins.bottom = 0.5

# Optional: center on page
sheet.page_setup.horizontalCentered = True


sheet.print_area = "A1:F85"
sheet.row_breaks.append(Break(id=66))
sheet.col_breaks.append(Break(id=6))

# Set default font for the workbook (matches Excel default)
default_font = Font(name="Helvetica", size=10)
for row in sheet.iter_rows(min_row=1, max_row=66, min_col=1, max_col=6):
    for cell in row:
        cell.font = default_font

for row_idx in range(1, 67):
    sheet.row_dimensions[row_idx].height = 12

sheet.column_dimensions["A"].width = 82 / 5.9
sheet.column_dimensions["B"].width = 266 / 5.9
sheet.column_dimensions["C"].width = 63 / 5.9
sheet.column_dimensions["D"].width = 63 / 5.9
sheet.column_dimensions["E"].width = 63 / 5.9
sheet.column_dimensions["F"].width = 63 / 5.9

# Text alignment in cells
alignment = Alignment(horizontal="left", vertical="center")

for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column
):
    for cell in row:
        cell.alignment = alignment

# Header
sheet['A1'] = 'Vyúčtování služební cesty'
sheet["A1"].font = Font(bold=True)

sheet['C1'] = 'Profisolv, s.r.o.'
sheet["C1"].font = Font(bold=True)

sheet['E1'] = 'Číslo:'
sheet["E1"].font = Font(bold=True)
sheet['E2'] = 'List:'
sheet["E2"].font = Font(bold=True)

sheet['A4'] = 'Pracovník:'
sheet["A4"].font = Font(bold=True)
sheet['A5'] = 'Ùčel cesty:'
sheet["A5"].font = Font(bold=True)
sheet['A6'] = 'Prostředek:'
sheet["A6"].font = Font(bold=True)
sheet['A7'] = 'Trasa:'
sheet["A7"].font = Font(bold=True)

# Route
sheet['B9'] = 'Popis trasy'
sheet["B9"].font = Font(bold=True)
sheet["B9"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

sheet['A10'] = 'Bod'
sheet["A10"].font = Font(bold=True)
sheet['B10'] = 'Místo'
sheet["B10"].font = Font(bold=True)
sheet['C10'] = 'Datum'
sheet["C10"].font = Font(bold=True)
sheet['D10'] = 'Čas'
sheet["D10"].font = Font(bold=True)
sheet['E10'] = 'Doba'
sheet["E10"].font = Font(bold=True)
sheet['F10'] = 'Jídla'
sheet["F10"].font = Font(bold=True)


# Costs
sheet['B32'] = 'Náklady'
sheet["B32"].font = Font(bold=True)
sheet["B32"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

sheet['A33'] = 'Stravné'
sheet["A33"].font = Font(bold=True)
sheet['A34'] = 'Datum'
sheet["A34"].font = Font(bold=True)
sheet['B34'] = 'Popis'
sheet["B34"].font = Font(bold=True)
sheet['E34'] = 'Plné'
sheet["E34"].font = Font(bold=True)
sheet['F34'] = 'Snížené'
sheet["F34"].font = Font(bold=True)
sheet['D42'] = 'Celkem:'
sheet['D43'] = 'Kapesné:'
sheet['F43'] = 'xxxxxxx'
sheet["F43"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

sheet['A45'] = 'Ubytování'
sheet["A45"].font = Font(bold=True)
sheet['A46'] = 'Datum'
sheet["A46"].font = Font(bold=True)
sheet['B46'] = 'Popis'
sheet["B46"].font = Font(bold=True)
sheet['E46'] = 'Doklad č.'
sheet["E46"].font = Font(bold=True)
sheet['F46'] = 'Částka'
sheet["F46"].font = Font(bold=True)
sheet['E52'] = 'Celkem:'

sheet['A54'] = 'Ostatní výdaje'
sheet["A54"].font = Font(bold=True)
sheet['A55'] = 'Datum'
sheet["A55"].font = Font(bold=True)
sheet['B55'] = 'Popis'
sheet["B55"].font = Font(bold=True)
sheet['E55'] = 'Doklad č.'
sheet["E55"].font = Font(bold=True)
sheet['F55'] = 'Částka'
sheet["F55"].font = Font(bold=True)
sheet['E62'] = 'Celkem:'

# Footer
sheet['A64'] = 'Zúčtováno dne:'
sheet['A65'] = 'Podpis'
sheet["E64"] = 'Záloha'
sheet['C65'] = 'Mezisoučet:'
sheet["C65"].font = Font(italic=True)
sheet["E65"] = 'Náklady'
sheet['E66'] = 'K vyplacení:'
sheet["E66"].font = Font(bold=True)

wb.save('../output/spam.xlsx')
